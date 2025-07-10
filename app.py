from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from flask import Flask, json, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from io import BytesIO, StringIO
from decimal import Decimal
import pandas as pd
import os
from urllib.parse import urlparse

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuração mais robusta para o banco de dados
# Arquivo: app.py
# Alterar aproximadamente nas linhas 25-37:

# Configuração para o banco de dados
database_url = os.environ.get('DATABASE_URL') or 'mysql://root:CDgpcTOfpvsefoWvuqPyZmrwnBjdfqjz@centerbeam.proxy.rlwy.net:10393/railway'

# Garantir que estamos usando o driver pymysql
if "pymysql" not in database_url:
    database_url = database_url.replace('mysql://', 'mysql+pymysql://')

app.config['SQLALCHEMY_DATABASE_URI'] = database_url

# Adicione estas configurações para melhorar a estabilidade
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 280,
    'pool_timeout': 20,
    'pool_pre_ping': True,
    'connect_args': {
        'connect_timeout': 30
    }
}

# Desative rastreamento de modificações para melhorar o desempenho
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

#Start URL
db = SQLAlchemy(app)

try:
    with app.app_context():
        db.engine.execute("SELECT 1")
    print("Conexão com o banco de dados estabelecida com sucesso!")
except Exception as e:
    print(f"Erro ao conectar ao banco de dados: {e}")

@app.route('/')
def select_farm():
    # Tela de seleção de fazendas será a página inicial
    return render_template('select_farm.html')

@app.route('/index/<int:farm_id>')
def index(farm_id):
    # Armazenar o ID da fazenda na sessão para uso em outras páginas
    session['farm_id'] = farm_id
    version = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Buscar o nome da fazenda para exibir no template
    query = text("SELECT nome FROM fazendas WHERE id = :farm_id")
    farm_name = db.session.execute(query, {'farm_id': farm_id}).scalar()
    
    return render_template('index.html', farm_name=farm_name)

@app.route('/consolidado')
def consolidado():
    # Página de relatório consolidado (a ser implementada futuramente)
    return render_template('consolidado.html')

# Certificar que a sessão seja inicializada
app.secret_key = 'frutosdovale_secret_key'  # Chave secreta para a sessão

@app.route('/tractors')
def tractors():
    # Obter farm_id da sessão
    farm_id = session.get('farm_id')
    # Se não existe farm_id na sessão, redireciona para a seleção de fazendas
    if farm_id is None:
        return redirect(url_for('select_farm'))
    
    return render_template('tractors.html', farm_id=farm_id)

@app.route('/funcionarios')
def funcionarios():
    # Obter farm_id da sessão
    farm_id = session.get('farm_id')
    # Se não existe farm_id na sessão, redireciona para a seleção de fazendas
    if farm_id is None:
        return redirect(url_for('select_farm'))
    
    return render_template('funcionarios.html', farm_id=farm_id)

@app.route('/apontamento')
def apontamento():
    # Obter farm_id da sessão
    farm_id = session.get('farm_id')
    # Se não existe farm_id na sessão, redireciona para a seleção de fazendas
    if farm_id is None:
        return redirect(url_for('select_farm'))
    
    return render_template('apontamento.html', farm_id=farm_id)

def style_excel_worksheet(ws):
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get the dimensions of the worksheet
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Style data cells and adjust column widths
    column_widths = []
    
    for col in range(1, max_col + 1):
        max_length = 0
        column = ws.column_dimensions[get_column_letter(col)]
        
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        column.width = adjusted_width

@app.route('/register_abastecimento', methods=['POST'])
def register_abastecimento():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            INSERT INTO abastecimento (farm_id, data, combustivel, tipo_trator, quantidade, horimetro)
            VALUES (:farm_id, :data, :combustivel, :tipo_trator, :quantidade, :horimetro)
        """)
        
        # Preparar os dados
        quantidade = float(request.form['quantidade'])
        tipo_trator = request.form['tipoTrator']
        
        if tipo_trator != 'POSTO DE COMBUSTÍVEL':
            quantidade = -quantidade
            
        data = {
            'farm_id': farm_id, 
            'data': datetime.strptime(request.form['data'], '%Y-%m-%d').date(),
            'combustivel': request.form['combustivel'],
            'tipo_trator': tipo_trator,
            'quantidade': quantidade,
            'horimetro': float(request.form['horimetro'])
        }
        
        # Executar a query
        db.session.execute(query, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Registro de abastecimento adicionado com sucesso!'
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao registrar abastecimento: {str(e)}'
        }), 400
    
@app.route('/register_manutencao', methods=['POST'])
def register_manutencao():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        # Criar query SQL para inserção
        query = text("""
            INSERT INTO manutencao_realizada (farm_id, data, tipo_trator, tipo_manutencao, horimetro, operador)
            VALUES (:farm_id, :data, :tipo_trator, :tipo_manutencao, :horimetro, :operador)
        """)
        
        # Preparar os dados
        data = {
            'farm_id': farm_id,
            'data': datetime.strptime(request.form['data'], '%Y-%m-%d').date(),
            'tipo_trator': request.form['tipoTrator'],
            'tipo_manutencao': request.form['tipoManutencao'],
            'horimetro': float(request.form['horimetro']),
            'operador': request.form['operador']
        }
        
        # Executar a query
        db.session.execute(query, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Registro de manutenção adicionado com sucesso!'
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao registrar manutenção: {str(e)}'
        }), 400
    
@app.route('/view_abastecimento', methods=['GET'])
def view_abastecimento():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            SELECT id, data, combustivel, tipo_trator, quantidade, horimetro 
            FROM abastecimento 
            WHERE farm_id = :farm_id
            ORDER BY data DESC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        logs = [{
            'id': row[0],
            'data': row[1].isoformat(),
            'combustivel': row[2],
            'tipo_trator': row[3],
            'quantidade': row[4],
            'horimetro': row[5]
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'logs': logs
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar registros de abastecimento: {str(e)}'
        }), 400
    
@app.route('/view_manutencao', methods=['GET'])
def view_manutencao():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            SELECT id, data, tipo_trator, tipo_manutencao, horimetro, operador 
            FROM manutencao_realizada 
            WHERE farm_id = :farm_id
            ORDER BY data DESC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        logs = [{
            'id': row[0],
            'data': row[1].isoformat(),
            'tipo_trator': row[2],
            'tipo_manutencao': row[3],
            'horimetro': row[4],
            'operador': row[5]
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'logs': logs
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar registros de manutenção: {str(e)}'
        }), 400

@app.route('/delete_abastecimento/<int:id>', methods=['DELETE'])
def delete_abastecimento(id):
    try:
        query = text("DELETE FROM abastecimento WHERE id = :id")
        db.session.execute(query, {'id': id})
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Registro excluído com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao excluir registro: {str(e)}'
        }), 400

@app.route('/delete_manutencao/<int:id>', methods=['DELETE'])
def delete_manutencao(id):
    try:
        query = text("DELETE FROM manutencao_realizada WHERE id = :id")
        db.session.execute(query, {'id': id})
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Registro excluído com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao excluir registro: {str(e)}'
        }), 400

@app.route('/check_maintenance', methods=['POST'])
def check_maintenance():
    try:
        data = request.get_json()
        current_horimeter = float(data['horimeter'])
        tractor_type = data['tractorType']
        
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        # Query para manutenções baseadas em horímetro
        horimeter_query = text("""
            WITH UltimaManutenção AS (
                SELECT 
                    tipo_manutencao,
                    MAX(horimetro) as ultimo_horimetro
                FROM manutencao_realizada
                WHERE tipo_trator = :tipo_trator
                AND farm_id = :farm_id
                GROUP BY tipo_manutencao
            )
            SELECT 
                cr.tipo_manutencao,
                cr.intervalo,
                COALESCE(um.ultimo_horimetro, 0) as ultimo_horimetro,
                m.nome as nome_maquina,
                CASE 
                    WHEN um.ultimo_horimetro IS NULL THEN cr.intervalo
                    ELSE (um.ultimo_horimetro + cr.intervalo)
                END as proximo_horimetro
            FROM consulta_revisao cr
            JOIN maquinas m ON m.id = cr.maquina_id
            LEFT JOIN UltimaManutenção um ON um.tipo_manutencao = cr.tipo_manutencao
            WHERE cr.maquina_id = (SELECT id FROM maquinas WHERE nome = :tipo_trator)
            AND cr.tipo_intervalo = 'HORIMETRO'
            AND (
                um.ultimo_horimetro IS NULL 
                OR (:current_horimeter - um.ultimo_horimetro) >= cr.intervalo
            )
        """)

        # Query para manutenções baseadas em dias
        days_query = text("""
            WITH UltimaManutenção AS (
                SELECT 
                    tipo_manutencao,
                    MAX(data) as ultima_data
                FROM manutencao_realizada
                WHERE tipo_trator = :tipo_trator
                AND farm_id = :farm_id
                GROUP BY tipo_manutencao
            )
            SELECT 
                cr.tipo_manutencao,
                cr.intervalo,
                COALESCE(um.ultima_data, DATE_SUB(CURRENT_DATE, INTERVAL cr.intervalo DAY)) as ultima_data,
                m.nome as nome_maquina
            FROM consulta_revisao cr
            JOIN maquinas m ON m.id = cr.maquina_id
            LEFT JOIN UltimaManutenção um ON um.tipo_manutencao = cr.tipo_manutencao
            WHERE cr.maquina_id = (SELECT id FROM maquinas WHERE nome = :tipo_trator)
            AND cr.tipo_intervalo = 'DIAS'
            HAVING DATEDIFF(CURRENT_DATE, ultima_data) >= cr.intervalo
        """)
        
        # Executa as queries
        horimeter_data = db.session.execute(horimeter_query, 
            {'tipo_trator': tractor_type, 'current_horimeter': current_horimeter, 'farm_id': farm_id}).fetchall()
        days_data = db.session.execute(days_query, 
            {'tipo_trator': tractor_type, 'farm_id': farm_id}).fetchall()
        
        tasks = []
        machine_name = None

        # Processa manutenções por horímetro
        for maintenance in horimeter_data:
            machine_name = maintenance.nome_maquina
            tasks.append({
                'maintenance': maintenance.tipo_manutencao,
                'horimeter': maintenance.proximo_horimetro,
                'difference': current_horimeter - maintenance.proximo_horimetro,
                'machine_name': maintenance.nome_maquina,
                'type': 'HORIMETRO'
            })

        # Processa manutenções por dias
        for maintenance in days_data:
            machine_name = maintenance.nome_maquina
            dias_passados = (datetime.now().date() - maintenance.ultima_data).days
            tasks.append({
                'maintenance': maintenance.tipo_manutencao,
                'days_passed': dias_passados,
                'interval': maintenance.intervalo,
                'last_date': maintenance.ultima_data.strftime('%d/%m/%Y'),
                'machine_name': maintenance.nome_maquina,
                'type': 'DIAS'
            })
        
        return jsonify({
            'status': 'success',
            'tasks': tasks,
            'machine_name': machine_name,
            'current_horimeter': current_horimeter
        })
        
    except Exception as e:
        print(f"Error in check_maintenance: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

@app.route('/download_excel/<table_name>')
def download_excel(table_name): 
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        wb = Workbook()
        ws = wb.active
        
        if table_name == 'abastecimento':
            # Query abastecimento data
            query = text("SELECT data, combustivel, tipo_trator, quantidade, horimetro FROM abastecimento WHERE farm_id = :farm_id ORDER BY data DESC")
            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
            
            # Handle date conversion with error handling
            def safe_date_format(date_str):
                try:
                    if pd.isna(date_str) or date_str == '1111-11-11':
                        return 'Data Inválida'
                    return pd.to_datetime(date_str).strftime('%d/%m/%Y')
                except:
                    return 'Data Inválida'
            
            # Apply safe date formatting
            df['Data'] = df['data'].apply(safe_date_format)
            
            # Rename columns and drop original date column
            df = df.rename(columns={
                'combustivel': 'Combustível',
                'tipo_trator': 'Tipo de Trator',
                'quantidade': 'Quantidade (L)',
                'horimetro': 'Horímetro'
            })
            df = df.drop('data', axis=1)
            
            # Reordenar as colunas
            df = df[['Data', 'Tipo de Trator', 'Combustível', 'Quantidade (L)', 'Horímetro']]
            
            ws.title = "Registro de Abastecimento"
            
        elif table_name == 'manutencao':
            # Query manutencao data
            query = text("SELECT data, tipo_trator, tipo_manutencao, horimetro, operador FROM manutencao_realizada ORDER BY data DESC")
            df = pd.read_sql(query, db.engine)
            
            # Apply same safe date formatting
            def safe_date_format(date_str):
                try:
                    if pd.isna(date_str) or date_str == '1111-11-11':
                        return 'Data Inválida'
                    return pd.to_datetime(date_str).strftime('%d/%m/%Y')
                except:
                    return 'Data Inválida'
            
            df['Data'] = df['data'].apply(safe_date_format)
            
            # Rename columns and drop original date column
            df = df.rename(columns={
                'tipo_trator': 'Tipo de Trator',
                'tipo_manutencao': 'Tipo de Manutenção',
                'horimetro': 'Horímetro',
                'operador': 'Operador'
            })
            df = df.drop('data', axis=1)
            
            # Reordenar as colunas
            df = df[['Data', 'Tipo de Trator', 'Tipo de Manutenção', 'Horímetro', 'Operador']]
            
            ws.title = "Registro de Manutenção"
            
        elif table_name == 'revisao':
        # Query revisao data
            query = text("""
                SELECT
                    m.nome as 'Trator',
                    cr.intervalo as 'Horímetro',
                    cr.tipo_manutencao as 'Serviço de Manutenção',
                    cr.tipo_intervalo as 'Tipo de Intervalo'
                FROM consulta_revisao cr
                JOIN maquinas m ON m.id = cr.maquina_id
                ORDER BY m.nome, cr.intervalo
            """)
            df = pd.read_sql(query, db.engine)

            # Não precisamos renomear as colunas pois já estão corretas da query
            # Reordenar as colunas mantendo o tipo de intervalo
            df = df[['Trator', 'Horímetro', 'Serviço de Manutenção', 'Tipo de Intervalo']]
            
            ws.title = "Cronograma de Revisões"
            
        else:
            return jsonify({'error': 'Tabela não encontrada'}), 404

        # Convert dataframe to excel rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format numeric cells (excluding the 'Data' column)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        # Apply styling
        style_excel_worksheet(ws)

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with current date
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"{ws.title}_{current_date}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"Error generating Excel file: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/get_maquinas', methods=['GET'])
def get_maquinas():
    try:
        farm_id = session.get('farm_id')
        
        query = text("""
            SELECT nome 
            FROM maquinas 
            WHERE ativo = 1 
            AND (farm_id = :farm_id OR farm_id IS NULL)
            ORDER BY nome
        """)
        
        maquinas = db.session.execute(query, {'farm_id': farm_id}).fetchall()
        
        return jsonify({
            'status': 'success',
            'maquinas': [maquina[0] for maquina in maquinas]
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

@app.route('/get_maintenance_types', methods=['GET'])
def get_maintenance_types():
    try:
        # Modificar a query para usar o campo correto
        query = text("""
            SELECT DISTINCT tipo_manutencao 
            FROM consulta_revisao 
            ORDER BY tipo_manutencao
        """)
        
        maintenance_types = db.session.execute(query).fetchall()
        
        return jsonify({
            'status': 'success',
            'types': [type[0] for type in maintenance_types]
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

@app.route('/get_funcionarios', methods=['GET'])
def get_funcionarios():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            SELECT id, nome 
            FROM funcionarios 
            WHERE ativo = 1 AND farm_id = :farm_id
            ORDER BY nome ASC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        funcionarios = [{
            'id': row[0],
            'nome': row[1]
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'funcionarios': funcionarios
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar funcionários: {str(e)}'
        }), 400

@app.route('/cadastrar_funcionario', methods=['POST'])
def cadastrar_funcionario():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        # Criar query SQL para inserção
        query = text("""
            INSERT INTO funcionarios 
            (farm_id, nome, sexo, data_admissao, cpf, pix, endereco, data_nascimento, 
            tipo_contratacao, funçao, empregador, sindicato, ativo, 
            ultimas_ferias, salario)
            VALUES 
            (:farm_id, :nome, :sexo, :data_admissao, :cpf, :pix, :endereco, :data_nascimento, 
            :tipo_contratacao, :funcao, 0, NULL, 1, :ultimas_ferias, :salario)
        """)

        # Dicionário de dados:
        data = {
            'farm_id': farm_id,
            'nome': request.form['nome'],
            'sexo': request.form['sexo'],
            'data_admissao': datetime.strptime(request.form['dataAdmissao'], '%Y-%m-%d').date(),
            'cpf': request.form['cpf'],
            'pix': request.form['pix'] if request.form['pix'] else None,
            'endereco': request.form['endereco'] if request.form['endereco'] else None,
            'data_nascimento': datetime.strptime(request.form['dataNascimento'], '%Y-%m-%d').date(),
            'tipo_contratacao': request.form['tipoContratacao'],
            'funcao': request.form['funcao'],
            'ultimas_ferias': datetime.strptime(request.form['ultimasFerias'], '%Y-%m-%d').date() if request.form.get('ultimasFerias') else None,
            'salario': float(request.form['salario']) if request.form.get('salario') else None
        }
        
        # Executar a query
        db.session.execute(query, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Funcionário cadastrado com sucesso!'
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao cadastrar funcionário: {str(e)}'
        }), 400

@app.route('/remover_funcionario/<int:id>', methods=['DELETE'])
def remover_funcionario(id):
    try:
        # Verificar registros em todas as tabelas relacionadas
        check_queries = [
            text("""
                SELECT COUNT(*) 
                FROM apontamento 
                WHERE funcionario_id = :id
            """),
            text("""
                SELECT COUNT(*)
                FROM registro_estoque
                WHERE funcionario_id = :id
            """)
        ]
        
        # Executar cada verificação
        for query in check_queries:
            result = db.session.execute(query, {'id': id}).scalar()
            if result > 0:
                return jsonify({
                    'status': 'error',
                    'message': 'Não é possível remover este funcionário pois existem registros vinculados a ele.'
                }), 400

        # Se não houver registros relacionados, proceder com a remoção
        query = text("DELETE FROM funcionarios WHERE id = :id")
        db.session.execute(query, {'id': id})
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Funcionário removido com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao remover funcionário: {str(e)}'
        }), 400

@app.route('/estoque')
def estoque():
    # Obter farm_id da sessão
    farm_id = session.get('farm_id')
    # Se não existe farm_id na sessão, redireciona para a seleção de fazendas
    if farm_id is None:
        return redirect(url_for('select_farm'))
    
    return render_template('estoque.html', farm_id=farm_id)

@app.route('/get_produtos', methods=['GET'])
def get_produtos():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            SELECT id, produto as nome, tipo, classificacao
            FROM produtos
            WHERE ativo = 1
            AND (farm_id = :farm_id OR farm_id IS NULL)
            ORDER BY produto ASC
        """)

        result = db.session.execute(query, {'farm_id': farm_id})
        produtos = [{
            'id': row[0],
            'nome': row[1],
            'tipo': row[2],
            'classificacao': row[3]
        } for row in result]

        return jsonify({
            'status': 'success',
            'produtos': produtos
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar produtos: {str(e)}'
        }), 400

@app.route('/get_lojas', methods=['GET'])
def get_lojas():
    try:
        query = text("""
            SELECT id, loja as nome
            FROM lojas 
            ORDER BY loja ASC
        """)
        
        result = db.session.execute(query)
        lojas = [{
            'id': row[0],
            'nome': row[1]
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'lojas': lojas
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar lojas: {str(e)}'
        }), 400

@app.route('/adicionar_produto', methods=['POST'])
def adicionar_produto():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            INSERT INTO produtos (farm_id, produto, tipo, classificacao, ativo)
            VALUES (:farm_id, :nome, :tipo, :classificacao, 1)
        """)

        db.session.execute(query, {
            'farm_id': farm_id,
            'nome': request.form['nome'],
            'tipo': request.form['tipo'],
            'classificacao': request.form['classificacao']
        })
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Produto adicionado com sucesso!'
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao adicionar produto: {str(e)}'
        }), 400

@app.route('/remover_produto/<int:id>', methods=['DELETE'])
def remover_produto(id):
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        query = text("""
            UPDATE produtos 
            SET ativo = 0 
            WHERE id = :id
            AND (farm_id = :farm_id OR farm_id IS NULL)
        """)
        
        result = db.session.execute(query, {'id': id, 'farm_id': farm_id})
        db.session.commit()
        
        if result.rowcount > 0:
            return jsonify({
                'status': 'success',
                'message': 'Produto desativado com sucesso!'
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Produto não encontrado ou não pertence a esta fazenda.'
            }), 404
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao desativar produto: {str(e)}'
        }), 400

@app.route('/registrar_estoque', methods=['POST'])
def registrar_estoque():
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        # Dados básicos que são comuns para entrada e saída
        data = {
            'farm_id': farm_id,
            'data': datetime.strptime(request.form['data'], '%Y-%m-%d').date(),
            'produto_id': request.form['produto_id'],
            'quantidade': float(request.form['quantidade']),
            'tipo_movimento': request.form['tipo_movimento'],
            'funcionario_id': request.form['funcionario_id'],
            'valor_unitario': None,
            'loja_id': None
        }
        
        if data['tipo_movimento'] == 'ENTRADA':
            # Adiciona os campos específicos para entrada
            data['loja_id'] = request.form['loja_id']
            data['valor_unitario'] = float(request.form['valor_unitario'])
        
        # Query para registro_estoque
        query_estoque = text("""
            INSERT INTO registro_estoque 
            (farm_id, data, produto_id, quantidade, tipo_movimento, valor_unitario, funcionario_id, loja_id)
            VALUES 
            (:farm_id, :data, :produto_id, :quantidade, :tipo_movimento, :valor_unitario, :funcionario_id, :loja_id)
        """)
        
        # Executar a query do registro_estoque
        db.session.execute(query_estoque, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Movimentação registrada com sucesso!'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao registrar movimentação: {str(e)}'
        }), 400
    
@app.route('/download_estoque/<tipo>', methods=['GET'])
def download_estoque(tipo): 
    try:
        # Obter farm_id da sessão
        farm_id = session.get('farm_id')
        
        wb = Workbook()
        ws = wb.active
        
        if tipo == 'movimentacoes':
            query = text("""
                SELECT 
                    re.data,
                    p.produto,
                    re.quantidade,
                    re.tipo_movimento,
                    COALESCE(re.valor_unitario, 0) as valor_unitario,
                    f.nome as funcionario,
                    COALESCE(l.loja, 'N/A') as loja
                FROM registro_estoque re
                JOIN produtos p ON re.produto_id = p.id
                JOIN funcionarios f ON re.funcionario_id = f.id
                LEFT JOIN lojas l ON re.loja_id = l.id
                WHERE re.farm_id = :farm_id
                ORDER BY re.data DESC
            """)
            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
            
            # Renomear colunas
            df = df.rename(columns={
                'data': 'Data',
                'produto': 'Produto',
                'quantidade': 'Quantidade',
                'tipo_movimento': 'Tipo de Movimento',
                'valor_unitario': 'Valor Unitário (R$)',
                'funcionario': 'Funcionário',
                'loja': 'Loja'
            })
            
            # Formatar data
            df['Data'] = pd.to_datetime(df['Data']).dt.strftime('%d/%m/%Y')
            
            ws.title = "Movimentações de Estoque"

        elif tipo == 'produtos':
            query = text("""
                SELECT 
                    produto as 'Nome do Produto',
                    tipo as 'Tipo',
                    classificacao as 'Classificação'
                FROM produtos 
                WHERE ativo = 1
                AND (farm_id = :farm_id OR farm_id IS NULL)
                ORDER BY produto
            """)
            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
            ws.title = "Lista de Produtos"

        elif tipo == 'resumo':
            query = text("""
                WITH UltimaEntrada AS (
                    SELECT 
                        re.produto_id,
                        re.valor_unitario,
                        l.loja,
                        re.data,
                        ROW_NUMBER() OVER (PARTITION BY re.produto_id ORDER BY re.data DESC, re.id DESC) as rn
                    FROM registro_estoque re
                    LEFT JOIN lojas l ON re.loja_id = l.id
                    WHERE re.tipo_movimento = 'ENTRADA'
                    AND re.farm_id = :farm_id
                )
                SELECT 
                    p.produto as 'Produto',
                    SUM(CASE 
                        WHEN re.tipo_movimento = 'ENTRADA' THEN re.quantidade 
                        WHEN re.tipo_movimento = 'SAIDA' THEN -re.quantidade 
                        ELSE 0 
                    END) as 'Saldo Atual',
                    ue.valor_unitario as 'Valor Unitário (R$)',
                    ue.loja as 'Última Loja'
                FROM produtos p
                LEFT JOIN registro_estoque re ON p.id = re.produto_id AND re.farm_id = :farm_id
                LEFT JOIN UltimaEntrada ue ON p.id = ue.produto_id AND ue.rn = 1
                WHERE p.ativo = 1
                AND (p.farm_id = :farm_id OR p.farm_id IS NULL)
                GROUP BY p.id, p.produto, ue.valor_unitario, ue.loja
                ORDER BY p.produto
            """)
            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
            
            # Calcular valor total por produto
            df['Valor Total'] = df['Saldo Atual'] * pd.to_numeric(df['Valor Unitário (R$)'].fillna(0))
            
            # Calcular o valor total do estoque
            valor_total_estoque = df['Valor Total'].sum()
            
            # Formatar valores
            df['Valor Unitário (R$)'] = df['Valor Unitário (R$)'].fillna(0).apply(lambda x: f'R$ {x:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'))
            df['Valor Total'] = df['Valor Total'].apply(lambda x: f'R$ {x:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'))
            
            # Adicionar linha do total
            df.loc[len(df.index)] = ['VALOR TOTAL DO ESTOQUE GALPÃO', '', '', '', f'R$ {valor_total_estoque:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')]
            
            ws.title = "Resumo do Estoque Galpão"

            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})

            # Garantir que as colunas numéricas sejam float
            df['Saldo Atual'] = pd.to_numeric(df['Saldo Atual'], errors='coerce').fillna(0)
            df['Valor Unitário (R$)'] = pd.to_numeric(df['Valor Unitário (R$)'], errors='coerce').fillna(0)
            df['Valor Total (R$)'] = pd.to_numeric(df['Valor Total (R$)'], errors='coerce').fillna(0)

            def safe_format_currency(value):
                try:
                    # Remover R$, pontos e substituir vírgula por ponto para conversão
                    numeric_value = float(str(value).replace('R$', '').replace('.', '').replace(',', '.').strip() or 0)
                    # Dividir por 100 para corrigir a representação de centavos
                    numeric_value = numeric_value / 100
                    return f'R$ {numeric_value:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                except (ValueError, TypeError):
                    return 'R$ 0,00'

            # Calcular o valor total do estoque
            valor_total_estoque = df['Valor Total (R$)'].sum()

            # Adicionar linha com o total geral
            total_row = pd.DataFrame([['VALOR TOTAL DO ESTOQUE', None, None, None, None, None, valor_total_estoque]], 
                                    columns=df.columns)
            df = pd.concat([df, total_row], ignore_index=True)

            # Formatar as colunas monetárias e numéricas
            for i, row in df.iterrows():
                if pd.notnull(row['Valor Unitário (R$)']):
                    df.at[i, 'Valor Unitário (R$)'] = safe_format_currency(row['Valor Unitário (R$)'])
                if pd.notnull(row['Valor Total (R$)']):
                    df.at[i, 'Valor Total (R$)'] = safe_format_currency(row['Valor Total (R$)'])
                if pd.notnull(row['Saldo Atual']):
                    df.at[i, 'Saldo Atual'] = f'{float(row["Saldo Atual"]):,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')

            ws.title = "Estoque Total"

            # Converter o DataFrame para Excel com tratamento de valores nulos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if pd.notnull(value) else '')

        style_excel_worksheet(ws)

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"estoque_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

    except Exception as e:
        print(f"Erro ao gerar arquivo Excel: {str(e)}")  # Para debug
        return jsonify({'error': str(e)}), 500

@app.route('/get_valvulas', methods=['GET'])
def get_valvulas():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            SELECT id, valvula, area_hectare, cavalo, variedade
            FROM valvulas 
            WHERE farm_id = :farm_id
            ORDER BY valvula ASC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        valvulas = [{
            'id': row[0],
            'valvula': row[1],
            'area_hectare': float(row[2]),
            'cavalo': row[3],
            'variedade': row[4]
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'valvulas': valvulas
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar válvulas: {str(e)}'
        }), 400

@app.route('/get_atividades', methods=['GET'])
def get_atividades():
    try:
        farm_id = session.get('farm_id')

        query = text("""
            SELECT id, atividade, valor_unit, meta
            FROM atividade
            WHERE ativo = 1
            AND (farm_id = :farm_id OR farm_id IS NULL)
            ORDER BY atividade ASC
        """)

        result = db.session.execute(query, {'farm_id': farm_id})

        atividades = [{
            'id': row[0],
            'atividade': row[1],
            'valor_unit': float(row[2]),
            'meta': float(row[3])
        } for row in result]

        return jsonify({
            'status': 'success',
            'atividades': atividades
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar atividades: {str(e)}'
        }), 400

# DESATIVAR loja (boa prática)
@app.route('/remover_loja/<int:id>', methods=['DELETE'])
def remover_loja(id):
    try:
        # Primeiro, verificar se existem registros vinculados no estoque
        check_estoque_query = text("""
            SELECT COUNT(*) 
            FROM registro_estoque 
            WHERE loja_id = :id
        """)
        estoque_count = db.session.execute(check_estoque_query, {'id': id}).scalar()
        
        if estoque_count > 0:
            return jsonify({
                'status': 'error',
                'message': 'Não é possível remover esta loja pois existem movimentações de estoque vinculadas a ela.'
            }), 400

        # Se não houver registros vinculados, proceder com a remoção
        delete_query = text("DELETE FROM lojas WHERE id = :id")
        result = db.session.execute(delete_query, {'id': id})
        
        if result.rowcount == 0:
            return jsonify({
                'status': 'error',
                'message': 'Loja não encontrada.'
            }), 404
            
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Loja removida com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao remover loja: {str(e)}'
        }), 400
    
@app.route('/adicionar_loja', methods=['POST'])
def adicionar_loja():
    try:
        nome_loja = request.form.get('nome')
        telefone = request.form.get('telefone')
        
        if not nome_loja:
            return jsonify({
                'status': 'error',
                'message': 'O nome da loja é obrigatório.'
            }), 400

        # Verificar se a loja já existe
        check_query = text("SELECT id FROM lojas WHERE loja = :nome")
        existing_loja = db.session.execute(check_query, {'nome': nome_loja}).first()
        
        if existing_loja:
            return jsonify({
                'status': 'error',
                'message': 'Uma loja com este nome já existe.'
            }), 400

        # Inserir nova loja
        query = text("""
            INSERT INTO lojas (loja, telefone) 
            VALUES (:nome, :telefone)
        """)
        
        db.session.execute(query, {
            'nome': nome_loja,
            'telefone': telefone
        })
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Loja adicionada com sucesso!'
        }), 201
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao adicionar loja: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro ao adicionar loja: {str(e)}'
        }), 400

@app.route('/registrar_apontamento', methods=['POST'])
def registrar_apontamento():
    try:
        farm_id = session.get('farm_id')
        # Validar dados recebidos
        if not all(field in request.form for field in ['funcionario_id', 'data', 'atividade_id']):
            return jsonify({
                'status': 'error',
                'message': 'Dados obrigatórios não fornecidos'
            }), 400

        # Obter dados do formulário com valores padrão seguros
        funcionario_id = request.form.get('funcionario_id')
        data = request.form.get('data')
        atividade_id = request.form.get('atividade_id')
        valvula = request.form.get('valvula')
        meta = Decimal(str(request.form.get('meta', 0)))
        realizado = Decimal(str(request.form.get('realizado', 0)))
        valor_unit = Decimal(str(request.form.get('valor_unit', 0)))
        tipo_apontamento = request.form.get('tipo_apontamento', 'Meta')  # Valor padrão 'Meta'
        observacao = request.form.get('observacao', '')
        hora = Decimal(str(request.form.get('hora', 0)))

        # Inicialização do extra
        extra = Decimal('0.00')

        # Cálculo do extra baseado no tipo de apontamento
        if tipo_apontamento == 'Meta':
            query = text("SELECT atividade FROM atividade WHERE id = :id")
            result = db.session.execute(query, {'id': atividade_id}).first()
            
            if result and result[0].upper() == 'FALTA':
                extra = Decimal('0.00')
            else:
                diferenca = realizado - meta
                if diferenca > 0:
                    extra = diferenca * valor_unit

        # IMPORTANTE: Incluir o tipo_apontamento na consulta SQL
        query = text("""
            INSERT INTO apontamento 
            (farm_id, funcionario_id, data, atividade_id, valvula_id, meta, realizado, 
             valor_unit, extra, observacao, hora, tipo_apontamento) 
            VALUES (:farm_id, :funcionario_id, :data, :atividade_id, :valvula_id, :meta, 
                    :realizado, :valor_unit, :extra, :observacao, :hora, :tipo_apontamento)
        """)
        
        valores = {
            'farm_id': farm_id,
            'funcionario_id': funcionario_id,
            'data': data,
            'atividade_id': atividade_id,
            'valvula_id': valvula,
            'meta': float(meta),
            'realizado': float(realizado),
            'valor_unit': float(valor_unit),
            'extra': float(extra),
            'observacao': observacao,
            'hora': float(hora),
            'tipo_apontamento': tipo_apontamento  # Adicionando o tipo_apontamento aos valores
        }
        
        db.session.execute(query, valores)
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Apontamento registrado com sucesso',
            'extra_calculado': str(extra)
        })

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao registrar apontamento: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    
@app.route('/get_ultimos_apontamentos', methods=['GET'])
def get_ultimos_apontamentos():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            SELECT 
                a.id,
                a.data,
                f.nome as funcionario,
                at.atividade,
                v.valvula,
                a.realizado,
                a.meta,
                a.valor_unit,
                a.extra,
                a.observacao
            FROM apontamento a
            JOIN funcionarios f ON a.funcionario_id = f.id
            JOIN atividade at ON a.atividade_id = at.id
            JOIN valvulas v ON a.valvula_id = v.id
            WHERE a.farm_id = :farm_id
            ORDER BY a.data DESC, a.id DESC
            LIMIT 10
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        apontamentos = [{
            'id': row.id,
            'funcionario': row.funcionario,
            'data': row.data.isoformat(),
            'atividade': row.atividade,
            'valvula': row.valvula,
            'realizado': float(row.realizado),
            'meta': float(row.meta),
            'valor_unit': float(row.valor_unit),
            'observacao': row.observacao
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'apontamentos': apontamentos
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar apontamentos: {str(e)}'
        }), 400

@app.route('/download_apontamento/<tipo>')
def download_apontamento(tipo):
    try:
        farm_id = session.get('farm_id')
        wb = Workbook()
        ws = wb.active
        
        if tipo == 'apontamentos':
            query = text("""
                SELECT 
                    a.data,
                    f.nome as funcionario,
                    at.atividade,
                    v.valvula,
                    a.meta,
                    a.realizado,
                    a.valor_unit,
                    a.extra,
                    a.observacao
                FROM apontamento a
                JOIN funcionarios f ON a.funcionario_id = f.id
                JOIN atividade at ON a.atividade_id = at.id
                JOIN valvulas v ON a.valvula_id = v.id
                WHERE a.farm_id = :farm_id
                ORDER BY a.data DESC
            """)
        
            df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
        
            if df.empty:
                return jsonify({'error': 'Não há dados de apontamentos para exportar'}), 404
        
            # Formatação das colunas
            df['Data'] = pd.to_datetime(df['data']).dt.strftime('%d/%m/%Y')
        
            # Renomear colunas
            df = df.rename(columns={
                'funcionario': 'Funcionário',
                'atividade': 'Atividade',
                'valvula': 'Válvula',
                'meta': 'Meta',
                'realizado': 'Realizado',
                'valor_unit': 'Valor Unitário',
                'extra': 'Extra',
                'observacao': 'Observação'
            })

            # Remover coluna data original e reorganizar colunas
            df = df.drop('data', axis=1)
            df = df[['Data', 'Funcionário', 'Atividade', 'Válvula', 'Meta', 'Realizado', 'Valor Unitário', 'Extra', 'Observação']]
        
            # Converter DataFrame para linhas do Excel
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Formatar valores numéricos
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
            
            ws.title = "Registro de Apontamentos"
            
        elif tipo == 'atividades':
            query = text("""
                SELECT 
                    atividade as 'Atividade',
                    CAST(valor_unit AS DECIMAL(10,2)) as 'Valor Unitário (R$)',
                    CAST(meta AS DECIMAL(10,2)) as 'Meta'
                FROM atividade 
                WHERE ativo = 1
                ORDER BY atividade
            """)
            
            # Executar a query e criar o DataFrame
            df = pd.read_sql(query, db.engine) 
            # Formatar as colunas numéricas
            df['Valor Unitário (R$)'] = df['Valor Unitário (R$)'].apply(lambda x: f'{float(x):,.2f}' if pd.notnull(x) else '')
            df['Meta'] = df['Meta'].apply(lambda x: f'{float(x):,.2f}' if pd.notnull(x) else '')
                
            # Configurar a planilha
            ws.title = "Lista de Atividades"
                
            # Adicionar cabeçalhos
            headers = list(df.columns)
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            # Adicionar dados
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                
    

        style_excel_worksheet(ws)

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"apontamento_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return jsonify({'error': str(e)}), 500
@app.route('/download_apontamento/resumo')
def download_resumo_apontamento():
    try:
        farm_id = session.get('farm_id')
        data_inicial = request.args.get('dataInicial')
        data_final = request.args.get('dataFinal')
        
        if not data_inicial or not data_final:
            return jsonify({'error': 'Datas não fornecidas'}), 400

        # Query para verificar se existem dados no período
        check_data_query = text("""
            SELECT COUNT(*) as total
            FROM apontamento
            WHERE data BETWEEN :data_inicial AND :data_final
            AND farm_id = :farm_id
        """)
        
        result = db.session.execute(check_data_query, {
            'data_inicial': data_inicial,
            'data_final': data_final,
            'farm_id': farm_id
        }).scalar()

        if result == 0:
            return jsonify({
                'error': 'Não foram encontrados registros para o período selecionado'
            }), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo Geral"

        # Query para o resumo geral
        query_resumo = text("""
            SELECT 
                f.nome as funcionario,
                f.pix as pix,
                f.tipo_contratacao as tipo_contratacao,
                COUNT(DISTINCT CASE 
                    WHEN at.atividade NOT IN ('FALTA', 'FOLGA', 'FÉRIAS') 
                    THEN a.data 
                    END) * CASE 
                        WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                        ELSE 0
                    END as salario_base,
                SUM(CASE WHEN a.tipo_apontamento IN ('Hora', 'Compensado') THEN a.hora ELSE 0 END) as horas,
                COALESCE(SUM(a.extra), 0) as extra_total,
                SUM(CASE WHEN at.atividade = 'FALTA' THEN 1 ELSE 0 END) as total_faltas,
                (COUNT(DISTINCT CASE 
                    WHEN at.atividade NOT IN ('FALTA', 'FOLGA', 'FÉRIAS') 
                    THEN a.data 
                    END) * CASE 
                        WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                        ELSE 0
                    END) + COALESCE(SUM(a.extra), 0) + 
                    CASE 
                        WHEN SUM(CASE WHEN a.tipo_apontamento IN ('Hora', 'Compensado') THEN 1 ELSE 0 END) > 0 
                        THEN COALESCE(SUM(
                            CASE 
                                WHEN a.tipo_apontamento IN ('Hora', 'Compensado') AND a.hora <= 0.5 THEN 5
                                WHEN a.tipo_apontamento IN ('Hora', 'Compensado') AND a.hora > 0.5 THEN 
                                    10 * FLOOR(a.hora) + CASE WHEN a.hora % 1 > 0 THEN 5 ELSE 0 END
                                ELSE 0
                            END
                        ), 0)
                        ELSE 0
                    END as total_periodo
            FROM apontamento a
            JOIN funcionarios f ON a.funcionario_id = f.id
            JOIN atividade at ON a.atividade_id = at.id
            WHERE a.data BETWEEN :data_inicial AND :data_final
            AND a.farm_id = :farm_id
            GROUP BY f.nome, f.pix, f.tipo_contratacao
            ORDER BY f.nome
        """)

        # Query para os detalhes individuais
        query_detalhes = text("""
            SELECT 
                f.nome as funcionario,
                a.data,
                at.atividade,
                a.meta,
                a.realizado,
                a.valor_unit,
                a.extra,
                a.hora,
                a.tipo_apontamento
            FROM apontamento a
            JOIN funcionarios f ON a.funcionario_id = f.id
            JOIN atividade at ON a.atividade_id = at.id
            WHERE a.data BETWEEN :data_inicial AND :data_final
            AND a.farm_id = :farm_id
            ORDER BY f.nome, a.data
        """)

        # Executar queries
        df_resumo = pd.read_sql(query_resumo, db.engine, params={
            'data_inicial': data_inicial,
            'data_final': data_final,
            'farm_id': farm_id
        })

        df_detalhes = pd.read_sql(query_detalhes, db.engine, params={
            'data_inicial': data_inicial,
            'data_final': data_final,
            'farm_id': farm_id
        })

        # Configurar aba de resumo geral
        headers = ['Funcionário', 'Salário Base (R$)', 'Extra (R$)', 'Total Período (R$)', 'PIX', 'Horas', 'Faltas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Preencher dados do resumo
        for idx, row in df_resumo.iterrows():
            r = idx + 2
            ws.cell(row=r, column=1, value=row['funcionario'])
            
            # Verificar tipo de contratação
            if row['tipo_contratacao'] == 'CLT':
                ws.cell(row=r, column=2, value='Salário CLT')
            else:
                ws.cell(row=r, column=2, value=float(row['salario_base'])).number_format = '#,##0.00'
            
            ws.cell(row=r, column=3, value=float(row['extra_total'])).number_format = '#,##0.00'
            
            # Verificar tipo de contratação para o total do período também
            if row['tipo_contratacao'] == 'CLT':
                valor_formatado = f'R$ {float(row["extra_total"]):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
                ws.cell(row=r, column=4, value=f'Salário CLT + {valor_formatado}')
            else:
                ws.cell(row=r, column=4, value=float(row['total_periodo'])).number_format = '#,##0.00'
            
            # Adicionar PIX, Horas e Faltas
            ws.cell(row=r, column=5, value=row['pix'] if pd.notna(row['pix']) else '')
            ws.cell(row=r, column=6, value=float(row['horas'])).number_format = '#,##0.00'
            ws.cell(row=r, column=7, value=int(row['total_faltas']))

        # Criar abas individuais para cada funcionário
        for funcionario in df_detalhes['funcionario'].unique():
            ws_func = wb.create_sheet(title=funcionario[:31])
            dados_func = df_detalhes[df_detalhes['funcionario'] == funcionario]
            
            headers = ['Data', 'Atividade', 'Meta', 'Realizado', 'Valor Unit.', 'Extra', 'Horas']
            for col, header in enumerate(headers, 1):
                cell = ws_func.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            
            # Para armazenar a soma de extras com bônus
            total_extra_com_bonus = 0
            
            for idx, row in dados_func.iterrows():
                r = idx - dados_func.index[0] + 2
                ws_func.cell(row=r, column=1, value=row['data'].strftime('%d/%m/%Y'))
                ws_func.cell(row=r, column=2, value=row['atividade'])
                ws_func.cell(row=r, column=3, value=float(row['meta'])).number_format = '#,##0.00'
                ws_func.cell(row=r, column=4, value=float(row['realizado'])).number_format = '#,##0.00'
                ws_func.cell(row=r, column=5, value=float(row['valor_unit'])).number_format = '#,##0.00'
                
                # Calcular extra com bônus de horas extras
                extra_original = float(row['extra'] or 0)
                horas_extras = float(row['hora'] or 0)
                bonus_horas = 0

                # Verificar se o tipo é 'Hora' ou 'Compensado' para adicionar bônus
                if str(row['tipo_apontamento']).strip() in ['Hora', 'Compensado'] and horas_extras > 0:
                    # Calcular o bônus baseado nas horas extras
                    if horas_extras <= 0.5:  # Até 30 minutos
                        bonus_horas = 5
                    else:  # Mais de 30 minutos
                        bonus_horas = 10 * int(horas_extras)  # R$10 por hora completa
                        if horas_extras % 1 > 0:  # Se tiver fração de hora
                            bonus_horas += 5  # Adiciona R$5
                    print(f"Calculando bônus para {row['funcionario']}: {bonus_horas} (tipo: {row['tipo_apontamento']}, horas: {horas_extras})")
                
                # Adicionar o bônus ao extra existente
                extra_total = extra_original + bonus_horas
                total_extra_com_bonus += extra_total
                
                # Atualizar o valor na célula de extra
                ws_func.cell(row=r, column=6, value=extra_total).number_format = '#,##0.00'
                ws_func.cell(row=r, column=7, value=float(row['hora'] or 0)).number_format = '#,##0.00'
            
            # Adicionar totais
            last_row = len(dados_func) + 2
            ws_func.cell(row=last_row, column=5, value='Totais:').font = Font(bold=True)
            
            # Usar o total calculado com bônus
            total_extra = ws_func.cell(row=last_row, column=6, value=total_extra_com_bonus)
            total_extra.font = Font(bold=True)
            total_extra.number_format = '#,##0.00'
            
            # Usar pandas para calcular o total de horas
            total_horas = ws_func.cell(row=last_row, column=7, value=float(dados_func['hora'].sum() or 0))
            total_horas.font = Font(bold=True)
            total_horas.number_format = '#,##0.00'
            
            style_excel_worksheet(ws_func)

        style_excel_worksheet(ws)

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        data_inicial_fmt = datetime.strptime(data_inicial, '%Y-%m-%d').strftime('%d_%m_%Y')
        data_final_fmt = datetime.strptime(data_final, '%Y-%m-%d').strftime('%d_%m_%Y')
        
        filename = f"apontamento_resumo_{data_inicial_fmt}_a_{data_final_fmt}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao gerar relatório: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/adicionar_atividade', methods=['POST'])
def adicionar_atividade():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            INSERT INTO atividade (atividade, valor_unit, meta, ativo, farm_id)
            VALUES (:atividade, :valor_unit, :meta, 1, :farm_id)
        """)
        
        data = {
            'atividade': request.form['atividade'],
            'valor_unit': float(request.form['valor_unit']),
            'meta': float(request.form['meta']) if request.form['meta'] else None,
            'farm_id': farm_id  # Adicionar esse parâmetro
        }
        
        db.session.execute(query, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Atividade adicionada com sucesso!'
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao adicionar atividade: {str(e)}'
        }), 400

@app.route('/desativar_atividade/<int:id>', methods=['PUT'])
def desativar_atividade(id):
    try:
        query = text("""
            UPDATE atividade
            SET ativo = 0
            WHERE id = :id
        """)

        result = db.session.execute(query, {'id': id})
        db.session.commit()

        if result.rowcount > 0:
            return jsonify({
                'status': 'success',
                'message': 'Atividade desativada com sucesso!'
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Atividade não encontrada.'
            }), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao desativar atividade: {str(e)}'
        }), 400

@app.route('/desativar_funcionario/<int:id>', methods=['PUT'])
def desativar_funcionario(id):
    try:
        query = text("""
            UPDATE funcionarios 
            SET ativo = 0 
            WHERE id = :id
        """)
        
        result = db.session.execute(query, {'id': id})
        db.session.commit()
        
        if result.rowcount > 0:
            return jsonify({
                'status': 'success',
                'message': 'Funcionário desativado com sucesso!'
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Funcionário não encontrado.'
            }), 404
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao desativar funcionário: {str(e)}'
        }), 400
    
@app.route('/download_funcionarios/<tipo>')
def download_funcionarios(tipo):
    try:
        wb = Workbook()
        ws = wb.active
        
        query = text("""
            SELECT 
                nome as "Nome",
                CASE 
                    WHEN sexo = 'M' THEN 'Masculino'
                    WHEN sexo = 'F' THEN 'Feminino'
                END as "Sexo",
                DATE_FORMAT(data_admissao, '%d/%m/%Y') as "Data Admissão",
                cpf as "CPF",
                CASE
                    WHEN empregador = 1 THEN 'Rogeria'
                    WHEN empregador = 2 THEN 'Adriana'
                    WHEN empregador = 3 THEN 'Cleomatson'
                    ELSE 'NÃO DEFINIDO'
                END as "Empregador",
                CASE
                    WHEN sindicato = 1 THEN 'SIM'
                    ELSE 'NÃO'
                END as "Sindicalizado",
                pix as "PIX",
                endereco as "Endereço",
                DATE_FORMAT(data_nascimento, '%d/%m/%Y') as "Data Nascimento",
                tipo_contratacao as "Tipo Contratação",
                funçao as "Função"
            FROM funcionarios
            WHERE ativo = 1
            ORDER BY nome ASC
        """)
        
        df = pd.read_sql(query, db.engine)
        ws.title = "Lista de Funcionários"
            
        # Converter dataframe para excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        style_excel_worksheet(ws)

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"funcionarios_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_total_combustivel/<combustivel>')
def get_total_combustivel(combustivel):
    try:
        query = text("""
            SELECT SUM(quantidade) as total
            FROM abastecimento
            WHERE combustivel = :combustivel
        """)
        
        result = db.session.execute(query, {'combustivel': combustivel}).first()
        
        return jsonify({
            'status': 'success',
            'total': float(result.total) if result.total else 0
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

@app.route('/get_ultimas_movimentacoes', methods=['GET'])
def get_ultimas_movimentacoes():
    try:
        farm_id = session.get('farm_id')
        page = request.args.get('page', 1, type=int)
        per_page = 5
        
        # Query para contar total de registros
        count_query = text("""
            SELECT COUNT(*) as total FROM registro_estoque
            WHERE farm_id = :farm_id  # Adicionar esta linha
        """)
        total_registros = db.session.execute(count_query, {'farm_id': farm_id}).scalar()
        
        # Query para buscar registros paginados
        query = text("""
            SELECT 
                re.id,
                re.data,
                p.produto,
                re.quantidade,
                re.tipo_movimento,
                COALESCE(re.valor_unitario, 0) as valor_unitario,
                f.nome as funcionario,
                COALESCE(l.loja, 'N/A') as loja
            FROM registro_estoque re
            JOIN produtos p ON re.produto_id = p.id
            JOIN funcionarios f ON re.funcionario_id = f.id
            LEFT JOIN lojas l ON re.loja_id = l.id
            WHERE re.farm_id = :farm_id  # Adicionar esta linha
            ORDER BY re.data DESC, re.id DESC
            LIMIT :limit OFFSET :offset
        """)
        
        offset = (page - 1) * per_page
        result = db.session.execute(query, {'limit': per_page, 'offset': offset, 'farm_id': farm_id})
        
        movimentacoes = [{
            'id': row.id,
            'data': row.data.strftime('%d/%m/%Y'),
            'produto': row.produto,
            'quantidade': float(row.quantidade),
            'tipo_movimento': row.tipo_movimento,
            'valor_unitario': float(row.valor_unitario),
            'funcionario': row.funcionario,
            'loja': row.loja
        } for row in result]
        
        total_pages = (total_registros + per_page - 1) // per_page
        
        return jsonify({
            'status': 'success',
            'movimentacoes': movimentacoes,
            'total_pages': total_pages,
            'current_page': page
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar movimentações: {str(e)}'
        }), 400

#Excluir um registro na lista(5) do estoque
@app.route('/excluir_movimentacao/<int:id>', methods=['DELETE'])
def excluir_movimentacao(id):
    try:
        # Verificar se o registro existe
        check_query = text("""
            SELECT id FROM registro_estoque WHERE id = :id
        """)
        result = db.session.execute(check_query, {'id': id}).first()
        
        if not result:
            return jsonify({
                'status': 'error',
                'message': 'Registro não encontrado.'
            }), 404
        
        # Excluir o registro do estoque
        query = text("DELETE FROM registro_estoque WHERE id = :id")
        db.session.execute(query, {'id': id})
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Registro excluído com sucesso!'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao excluir registro: {str(e)}'
        }), 400

@app.route('/excluir_apontamento/<int:id>', methods=['DELETE'])
def excluir_apontamento(id):
    try:
        query = text("DELETE FROM apontamento WHERE id = :id")
        db.session.execute(query, {'id': id})
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Apontamento excluído com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao excluir apontamento: {str(e)}'
        }), 400

@app.route('/get_ultimos_apontamentos_paginado', methods=['GET'])
def get_ultimos_apontamentos_paginado():
    try:
        farm_id = session.get('farm_id')
        page = request.args.get('page', 1, type=int)
        per_page = 7
        
        # Query para contar total de registros
        count_query = text("""
            SELECT COUNT(*) FROM apontamento
            WHERE farm_id = :farm_id  # Adicionar esta linha
        """)
        total_registros = db.session.execute(count_query, {'farm_id': farm_id}).scalar()
        
        # Calcular o offset baseado na página atual
        offset = (page - 1) * per_page
        
        query = text("""
            SELECT 
                a.id,
                a.data,
                f.nome as funcionario,
                at.atividade,
                v.valvula,
                a.realizado,
                a.meta,
                a.valor_unit,
                a.extra,
                a.hora,
                a.observacao
            FROM apontamento a
            JOIN funcionarios f ON a.funcionario_id = f.id
            JOIN atividade at ON a.atividade_id = at.id
            JOIN valvulas v ON a.valvula_id = v.id
            WHERE a.farm_id = :farm_id  # Adicionar esta linha
            ORDER BY a.id DESC
            LIMIT :limit OFFSET :offset
        """)
        
        result = db.session.execute(query, {'limit': per_page, 'offset': offset, 'farm_id': farm_id})
        apontamentos = [{
            'id': row.id,
            'data': row.data.isoformat(),
            'funcionario': row.funcionario,
            'atividade': row.atividade,
            'realizado': float(row.realizado),
            'meta': float(row.meta),
            'extra': float(row.extra) if row.extra else 0,
            'hora': float(row.hora) if row.hora else 0
        } for row in result]
        
        total_pages = (total_registros + per_page - 1) // per_page
        
        return jsonify({
            'status': 'success',
            'apontamentos': apontamentos,
            'current_page': page,
            'total_pages': total_pages
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar apontamentos: {str(e)}'
        }), 400

@app.route('/alterar_funcionario/<int:id>', methods=['PUT'])
def alterar_funcionario(id):
    try:
        query = text("""
            UPDATE funcionarios 
            SET nome = :nome,
                sexo = :sexo,
                data_admissao = :data_admissao,
                cpf = :cpf,
                pix = :pix,
                endereco = :endereco,
                data_nascimento = :data_nascimento,
                tipo_contratacao = :tipo_contratacao,
                funçao = :funcao,
                ultimas_ferias = :ultimas_ferias,
                salario = :salario
            WHERE id = :id
        """)
        
        data = {
            'id': id,
            'nome': request.form['nome'],
            'sexo': request.form['sexo'],
            'data_admissao': datetime.strptime(request.form['dataAdmissao'], '%Y-%m-%d').date(),
            'cpf': request.form['cpf'],
            'pix': request.form['pix'] if request.form['pix'] else None,
            'endereco': request.form['endereco'] if request.form['endereco'] else None,
            'data_nascimento': datetime.strptime(request.form['dataNascimento'], '%Y-%m-%d').date() if request.form['dataNascimento'] else None,
            'tipo_contratacao': request.form['tipoContratacao'],
            'funcao': request.form['funcao'],
            'ultimas_ferias': datetime.strptime(request.form['ultimasFerias'], '%Y-%m-%d').date() if request.form.get('ultimasFerias') else None,
            'salario': float(request.form['salario']) if request.form.get('salario') else None
        }
        
        db.session.execute(query, data)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Funcionário atualizado com sucesso!'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao atualizar funcionário: {str(e)}'
        }), 400

@app.route('/get_funcionario/<int:id>')
def get_funcionario(id):
    try:
        query = text("""
            SELECT 
                nome, sexo, data_admissao, cpf, pix, 
                endereco, data_nascimento, tipo_contratacao, funçao,
                ultimas_ferias, salario
            FROM funcionarios 
            WHERE id = :id AND ativo = 1
        """)
        
        result = db.session.execute(query, {'id': id}).first()
        
        if result:
            funcionario = {
                'nome': result.nome,
                'sexo': result.sexo,
                'cpf': result.cpf,
                'data_admissao': result.data_admissao if isinstance(result.data_admissao, str) else result.data_admissao.strftime('%Y-%m-%d') if result.data_admissao else None,
                'data_nascimento': result.data_nascimento if isinstance(result.data_nascimento, str) else result.data_nascimento.strftime('%Y-%m-%d') if result.data_nascimento else None,
                'tipo_contratacao': result.tipo_contratacao,
                'funcao': result.funçao,
                'pix': result.pix if result.pix else '',
                'endereco': result.endereco if result.endereco else ''
            }
            
            return jsonify({
                'status': 'success',
                'funcionario': funcionario
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Funcionário não encontrado'
            }), 404
            
    except Exception as e:
        print(f"Erro detalhado: {str(e)}")  # Log para debug
        return jsonify({
            'status': 'error',
            'message': 'Erro ao buscar dados do funcionário'
        }), 400

    
@app.route('/relatorios')
def relatorios():
    # Obter farm_id da sessão
    farm_id = session.get('farm_id')
    # Se não existe farm_id na sessão, redireciona para a seleção de fazendas
    if farm_id is None:
        return redirect(url_for('select_farm'))
    
    return render_template('relatorios.html', farm_id=farm_id)

@app.route('/get_report_data')
def get_report_data():
    try:
        farm_id = session.get('farm_id')
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        
        # Query para métricas gerais com tratamento para valores nulos
        metrics_query = text("""
        WITH CostData AS (
            SELECT 
                (SELECT COALESCE(SUM(
                    CASE 
                        WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                        ELSE 0 
                    END + COALESCE(extra, 0)), 0)
                FROM apontamento a
                JOIN valvulas v ON a.valvula_id = v.id
                JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE data BETWEEN :start_date AND :end_date
                AND v.id != 33
                AND a.farm_id = :farm_id) as labor_costs  # Adicionar farm_id
            ),
            FuelData AS (
                SELECT 
                    tipo_trator,
                    ABS(SUM(CASE 
                        WHEN quantidade IS NOT NULL THEN quantidade 
                        ELSE 0 
                    END)) as consumo
                FROM abastecimento
                WHERE data BETWEEN :start_date AND :end_date
                AND tipo_trator != 'POSTO DE COMBUSTÍVEL'
                AND farm_id = :farm_id
                GROUP BY tipo_trator
                HAVING consumo > 0
            ),
            EmployeeData AS (
                SELECT COUNT(*) as active_employees
                FROM funcionarios
                WHERE ativo = 1
                AND farm_id = :farm_id
            )
            SELECT 
                cd.labor_costs as total_costs,
                cd.labor_costs,
                (
                    SELECT JSON_ARRAYAGG(
                        JSON_OBJECT(
                            'maquina', fd.tipo_trator,
                            'consumo', fd.consumo
                        )
                    )
                    FROM FuelData fd
                ) as fuel_data,
                ad.total_applications,
                ed.active_employees
            FROM CostData cd
            CROSS JOIN ApplicationData ad
            CROSS JOIN EmployeeData ed
        """)

        # E atualizar o processamento do resultado:
        metrics_result = db.session.execute(metrics_query, {
            'start_date': start_date,
            'end_date': end_date,
            'farm_id': farm_id
        }).first()

        # Tratar dados do combustível
        fuel_data = json.loads(metrics_result.fuel_data) if metrics_result.fuel_data else []
        
        # Query para custos por válvula com tratamento para valores nulos
        valve_costs_query = text("""
            WITH ValveCosts AS (
                SELECT 
                    v.valvula,
                    v.id,
                    COALESCE(v.area_hectare, 0) as area_hectare,
                    -- Custos de mão de obra (apontamentos)
                    COALESCE(SUM(
                        CASE WHEN a.data BETWEEN :start_date AND :end_date 
                        THEN 
                            CASE 
                                WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                                ELSE 0 
                            END + COALESCE(a.extra, 0)
                        ELSE 0 END
                    ), 0) as labor_cost,
                    -- Custos de insumos
                    COALESCE((
                        SELECT SUM(
                            c.quantidade * COALESCE(
                                (SELECT valor_unitario 
                                FROM registro_estoque re 
                                WHERE re.produto_id = c.produto_id 
                                AND re.tipo_movimento = 'ENTRADA'
                                AND re.data <= c.data
                                ORDER BY re.data DESC, re.id DESC
                                LIMIT 1),
                                0
                            ) / 1000
                        )
                        FROM calda c
                        WHERE a.setor = v.id
                        AND a.realizado = 1
                        AND a.data BETWEEN :start_date AND :end_date
                        AND a.farm_id = :farm_id
                        AND c.tipo_movimento = 'SAIDA'
                    ), 0) as inputs_cost
                FROM valvulas v
                LEFT JOIN apontamento a ON v.id = a.valvula_id
                LEFT JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE v.id != 33
                AND v.farm_id = :farm_id
                GROUP BY v.id, v.valvula, v.area_hectare
            )
            SELECT 
                valvula,
                area_hectare,
                labor_cost,
                inputs_cost,
                (labor_cost + inputs_cost) as total_cost,
                CASE 
                    WHEN area_hectare > 0 THEN (labor_cost + inputs_cost) / area_hectare
                    ELSE 0
                END as cost_per_hectare
            FROM ValveCosts
            ORDER BY 
                CASE 
                    WHEN valvula = 'GERAL' THEN 0 
                    ELSE 1 
                END,
                valvula
        """)

        valve_costs = db.session.execute(valve_costs_query, {
            'start_date': start_date,
            'end_date': end_date,
            'farm_id': farm_id
        }).fetchall()
        
        
        # Formatar resposta
        return jsonify({
            'status': 'success',
            'metrics': {
                'totalCosts': float(metrics_result.total_costs or 0),
                'fuel_data': fuel_data, 
                'totalApplications': metrics_result.total_applications,
                'activeEmployees': metrics_result.active_employees
            },
            'tables': [{
                'valve': row.valvula,
                'laborCost': float(row.labor_cost or 0),
                'inputsCost': float(row.inputs_cost or 0),
                'totalCost': float(row.total_cost or 0),
                'costPerHectare': float(row.cost_per_hectare or 0)
            } for row in valve_costs],
            'kpis': calculate_kpis(metrics_result, valve_costs)
        })
        
    except Exception as e:
        print(f"Erro detalhado: {str(e)}")  # Log para debug
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

def calculate_kpis(metrics, valve_costs):
    try:
        # Área total fixa em 10 hectares
        AREA_TOTAL = 24.7
        
        # Calcular custos totais
        total_costs = float(metrics.total_costs or 0)
        
        # Custo médio por hectare
        avg_cost_per_hectare = total_costs / AREA_TOTAL if AREA_TOTAL > 0 else 0
        
        # Calcular consumo total de combustível (melhorado)
        total_fuel = 0
        if metrics.fuel_data:
            fuel_data = json.loads(metrics.fuel_data)
            for item in fuel_data:
                if isinstance(item, dict) and 'consumo' in item:
                    total_fuel += float(item['consumo'])
        
        # Consumo médio de combustível por hectare (melhorado)
        avg_fuel_consumption = total_fuel / AREA_TOTAL if AREA_TOTAL > 0 else 0
        
        return {
            'avgCostPerHectare': round(avg_cost_per_hectare, 2),
            'avgFuelConsumption': round(avg_fuel_consumption, 2),
            'totalArea': AREA_TOTAL,
            'totalFuelConsumption': round(total_fuel, 2)
        }
    except Exception as e:
        print(f"Erro no cálculo de KPIs: {str(e)}")
        return {
            'avgCostPerHectare': 0,
            'avgFuelConsumption': 0,
            'totalArea': AREA_TOTAL,
            'totalFuelConsumption': 0
        }
    
@app.route('/get_vendas_uvas')
def get_vendas_uvas():
    try:
        data_inicial = request.args.get('data_inicial')
        data_final = request.args.get('data_final')
        
        query = text("""
            SELECT 
                DATE_FORMAT(data, '%m/%Y') as mes,
                SUM(qte * 5) as quilo
            FROM vendas_uva
            WHERE data BETWEEN :data_inicial AND :data_final
            GROUP BY mes
            ORDER BY STR_TO_DATE(mes, '%m/%Y')
        """)
        
        result = db.session.execute(query, {
            'data_inicial': data_inicial,
            'data_final': data_final
        })
        
        vendas = [{
            'mes': row.mes,
            'quilo': float(row.quilo)
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'vendas': vendas
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400
    
@app.route('/download_vendas_excel')
def download_vendas_excel():
    try:
        farm_id = session.get('farm_id')
        
        # Definir as funções de formatação primeiro
        def format_currency(value):
            if pd.isna(value):
                return 'R$ 0,00'
            return f'R$ {float(value):,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')

        def format_kg(value):
            if pd.isna(value):
                return '0 kg'
            return f'{int(value):,} kg'.replace(',', '.')

        def format_integer(value):
            if pd.isna(value):
                return '0'
            return f'{int(value):,}'.replace(',', '.')

        ano = request.args.get('ano', '2025')
        
        wb = Workbook()
        ws_relatorio = wb.active  # Primeira aba
        ws_relatorio.title = "Relatório"
        
        # CRIAR AS OUTRAS ABAS NECESSÁRIAS
        ws_vendas = wb.create_sheet(title="Vendas")
        ws_produtos = wb.create_sheet(title="Produtos")

        # Query vendas - MODIFICADA PARA REMOVER A CONDIÇÃO FARM_ID
        query_vendas = text("""
            SELECT 
                vu.data as 'Data',
                vu.codigo as 'Documento',
                vu.area as 'Setor',
                vu.produto as 'Produto',
                vu.qte as 'Quantidade',
                CAST(vu.valor_pago/vu.qte AS DECIMAL(10,2)) as 'Valor Unitário',
                vu.valor_pago as 'Valor Total',
                vu.qte * 5 as 'Peso (kg)',
                CONCAT('Pallet: ', vu.no_pallet) as 'Observação',
                vu.semana as 'Semana'
            FROM vendas_uva vu
            WHERE YEAR(vu.data) = :ano
            ORDER BY vu.data DESC
        """)
        
        df_vendas = pd.read_sql(query_vendas, db.engine, params={'ano': ano})
        df_vendas['Data'] = pd.to_datetime(df_vendas['Data']).dt.strftime('%d/%m/%Y')
        
        # Formatação das colunas de vendas
        df_vendas['Valor Unitário'] = df_vendas['Valor Unitário'].apply(format_currency)
        df_vendas['Valor Total'] = df_vendas['Valor Total'].apply(format_currency)
        df_vendas['Peso (kg)'] = df_vendas['Peso (kg)'].apply(format_kg)
        df_vendas['Quantidade'] = df_vendas['Quantidade'].apply(format_integer)
        
        # Adicionar dados à aba de vendas
        for r_idx, row in enumerate(dataframe_to_rows(df_vendas, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws_vendas.cell(row=r_idx, column=c_idx, value=value)

        # Query resumo - MODIFICADA PARA REMOVER A CONDIÇÃO FARM_ID
        query_resumo = text("""
            SELECT 
                produto as 'Produto',
                COUNT(*) as 'Total Vendas',
                SUM(qte) as 'Total Caixas',
                SUM(qte * 5) as 'Total Kg',
                AVG(valor_pago / qte) as 'Preço Médio',
                SUM(valor_pago) as 'Valor Total'
            FROM vendas_uva
            WHERE YEAR(data) = :ano
            GROUP BY produto
            ORDER BY SUM(qte * 5) DESC
        """)

        df_relatorio = pd.read_sql(query_resumo, db.engine, params={'ano': ano})
        
        # Formatação das colunas do relatório
        df_relatorio['Preço Médio'] = df_relatorio['Preço Médio'].apply(format_currency)
        df_relatorio['Valor Total'] = df_relatorio['Valor Total'].apply(format_currency)
        df_relatorio['Total Kg'] = df_relatorio['Total Kg'].apply(format_kg)
        df_relatorio['Total Caixas'] = df_relatorio['Total Caixas'].apply(format_integer)
        df_relatorio['Total Vendas'] = df_relatorio['Total Vendas'].apply(format_integer)
        
        # Adicionar dados à aba de relatório
        for r_idx, row in enumerate(dataframe_to_rows(df_relatorio, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws_relatorio.cell(row=r_idx, column=c_idx, value=value)

        # Query produtos - MODIFICADA PARA REMOVER A CONDIÇÃO FARM_ID
        query_produtos = text("""
            SELECT 
                DISTINCT produto as 'Produto',
                '5 kg' as 'Tipo Embalagem'
            FROM vendas_uva
            WHERE YEAR(data) = :ano
            ORDER BY produto
        """)
        
        df_produtos = pd.read_sql(query_produtos, db.engine, params={'ano': ano})
        
        for r_idx, row in enumerate(dataframe_to_rows(df_produtos, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws_produtos.cell(row=r_idx, column=c_idx, value=value)

        # Aplicar estilos em todas as abas
        for ws in [ws_relatorio, ws_vendas, ws_produtos]:
            style_excel_worksheet(ws)

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'vendas_uvas_{ano}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        print(f"Erro ao gerar Excel de vendas: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_low_stock_products')
def get_low_stock_products():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            WITH EstoqueAtual AS (
                SELECT 
                    produto_id,
                    SUM(CASE 
                        WHEN tipo_movimento = 'ENTRADA' THEN quantidade 
                        WHEN tipo_movimento = 'SAIDA' THEN -quantidade 
                        ELSE 0 
                    END) as saldo_atual
                FROM registro_estoque
                WHERE farm_id = :farm_id  # Adicionar esta linha
                GROUP BY produto_id
            )
            SELECT 
                p.produto,
                p.tipo,
                p.classificacao,
                COALESCE(ea.saldo_atual, 0) as saldo_atual
            FROM produtos p
            LEFT JOIN EstoqueAtual ea ON p.id = ea.produto_id
            WHERE p.ativo = 1
            AND (p.farm_id = :farm_id OR p.farm_id IS NULL)  # Adicionar esta linha
            AND COALESCE(ea.saldo_atual, 0) < 5
            ORDER BY ea.saldo_atual ASC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        produtos = [{
            'produto': row.produto,
            'tipo': row.tipo,
            'classificacao': row.classificacao,
            'saldo_atual': float(row.saldo_atual)
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'produtos': produtos
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar produtos com baixo estoque: {str(e)}'
        }), 400
    
@app.route('/get_total_stock', methods=['GET'])
def get_total_stock():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            WITH UltimoValor AS (
                SELECT 
                    produto_id,
                    MAX(id) as ultimo_id
                FROM registro_estoque
                WHERE tipo_movimento = 'ENTRADA'
                AND valor_unitario IS NOT NULL
                AND farm_id = :farm_id  # Adicionar esta linha
                GROUP BY produto_id
            ),
            EstoqueAtual AS (
                SELECT 
                    p.id,
                    SUM(CASE 
                        WHEN re.tipo_movimento = 'ENTRADA' THEN re.quantidade 
                        WHEN re.tipo_movimento = 'SAIDA' THEN -re.quantidade 
                        ELSE 0 
                    END) as saldo_atual
                FROM produtos p
                LEFT JOIN registro_estoque re ON p.id = re.produto_id
                WHERE p.ativo = 1
                AND (p.farm_id = :farm_id OR p.farm_id IS NULL)  # Adicionar esta linha
                AND (re.farm_id = :farm_id OR re.farm_id IS NULL)  # Adicionar esta linha
                GROUP BY p.id
                HAVING saldo_atual > 0
            )
            SELECT SUM(
                ea.saldo_atual * COALESCE(
                    (SELECT valor_unitario 
                     FROM registro_estoque re2 
                     WHERE re2.id = uv.ultimo_id),
                    0
                )
            ) as valor_total
            FROM EstoqueAtual ea
            LEFT JOIN UltimoValor uv ON ea.id = uv.produto_id
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id}).scalar()
        
        return jsonify({
            'status': 'success',
            'valor_total': float(result) if result else 0
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400

@app.route('/download_funcionarios/<tipo>')
def download_funcionarios_relatório(tipo):
    try:
        wb = Workbook()
        ws = wb.active
        
        query = text("""
            SELECT 
                nome as "Nome",
                CASE 
                    WHEN sexo = 'M' THEN 'Masculino'
                    WHEN sexo = 'F' THEN 'Feminino'
                END as "Sexo",
                DATE_FORMAT(data_admissao, '%d/%m/%Y') as "Data Admissão",
                cpf as "CPF",
                CASE
                    WHEN empregador = 1 THEN 'Rogeria'
                    WHEN empregador = 2 THEN 'Adriana'
                    WHEN empregador = 3 THEN 'Cleomatson'
                    ELSE 'NÃO DEFINIDO'
                END as "Empregador",
                CASE
                    WHEN sindicato = 1 THEN 'SIM'
                    ELSE 'NÃO'
                END as "Sindicalizado",
                pix as "PIX",
                endereco as "Endereço",
                DATE_FORMAT(data_nascimento, '%d/%m/%Y') as "Data Nascimento",
                DATE_FORMAT(ultimas_ferias, '%m/%Y') as "Últimas Férias",
                CONCAT('R$ ', FORMAT(salario, 2, 'pt_BR')) as "Salário",
                tipo_contratacao as "Tipo Contratação",
                funçao as "Função"
            FROM funcionarios
            WHERE ativo = 1
            ORDER BY nome ASC
        """)
        
        df = pd.read_sql(query, db.engine)
        ws.title = "Lista de Funcionários"
            
        # Converter dataframe para excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        style_excel_worksheet(ws)

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Salvar arquivo
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"funcionarios_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/importar_vendas', methods=['POST'])
def importar_vendas():
    farm_id = session.get('farm_id')
    try:
        file = request.files['file']
        if not file:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400

        # Ler o conteúdo do arquivo
        content = file.read().decode('utf-8')
        print("Conteúdo do arquivo lido com sucesso")
        
        # Usar o pandas para ler o CSV
        df = pd.read_csv(StringIO(content), delimiter=';')
        print(f"DataFrame criado com sucesso. Número de linhas: {len(df)}")
        print("Colunas encontradas:", df.columns.tolist())
        
        # Processar cada linha
        registros_inseridos = 0
        for index, row in df.iterrows():
            try:
                print(f"\nProcessando linha {index + 1}:")
                print(row)
                
                # Verificar se a data está presente e no formato correto
                data_str = str(row['Data']).strip()
                print(f"Data encontrada: {data_str}")
                
                try:
                    data_obj = datetime.strptime(data_str, '%d/%m/%Y')
                except ValueError as e:
                    print(f"Erro ao converter data '{data_str}': {e}")
                    continue

                query = text("""
                    INSERT INTO vendas_uva 
                    (farm_id, cooperado, codigo, produto, no_pallet, semana, data, qte, area, valor_pago)
                    VALUES 
                    (:farm_id, :cooperado, :codigo, :produto, :no_pallet, :semana, :data, :qte, :area, :valor_pago)
                """)

                data = {
                    'farm_id': farm_id,  
                    'cooperado': str(row['Cooperado']),
                    'codigo': str(row['Codigo']),
                    'produto': str(row['Produto']),
                    'no_pallet': str(row['No. Pallet']),
                    'semana': int(row['Semana']),
                    'data': data_obj.date(),
                    'qte': int(row['Caixas']),
                    'area': float(row['AREA']),
                    'valor_pago': float(row['valor pago'])
                }
                
                print("Dados preparados:", data)
                
                db.session.execute(query, data)
                registros_inseridos += 1
                print(f"Registro {registros_inseridos} inserido com sucesso")
                
            except Exception as e:
                print(f"Erro ao processar linha {index + 1}: {str(e)}")
                print("Dados da linha:", row)
                continue
        
        db.session.commit()
        print(f"Total de registros inseridos: {registros_inseridos}")
        
        return jsonify({
            'message': f'Dados importados com sucesso. {registros_inseridos} registros inseridos.',
            'registros': registros_inseridos
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro geral na importação: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/processar_vendas_uva', methods=['POST'])
def processar_vendas_uva():
    try:
        farm_id = session.get('farm_id')
        file = request.files['file']
        df = pd.read_csv(file, encoding='utf-8')
        
        # Renomear colunas para corresponder ao banco de dados
        df = df.rename(columns={
            'Cooperado': 'cooperado',
            'Codigo': 'codigo',
            'Produto': 'produto',
            'No. Pallet': 'no_pallet',
            'Semana': 'semana',
            'Data': 'data',
            'Caixas': 'qte',
            'AREA': 'area',
            'valor pago': 'valor_pago'
        })
        
        # Processar data
        df['data'] = pd.to_datetime(df['data'], format='%d/%m/%Y')
        
        # Processar valor_pago (remover R$ e converter para float)
        df['valor_pago'] = df['valor_pago'].str.replace('R$', '').str.replace('.', '').str.replace(',', '.').astype(float)
        
        # Garantir que area seja string
        df['area'] = df['area'].astype(str)
        
        # Inserir dados no banco de dados
        for _, row in df.iterrows():
            query = text("""
                INSERT INTO vendas_uva 
                (farm_id, cooperado, codigo, produto, no_pallet, semana, data, qte, area, valor_pago)
                VALUES 
                (:farm_id, :cooperado, :codigo, :produto, :no_pallet, :semana, :data, :qte, :area, :valor_pago)
            """)
            
            db.session.execute(query, {
                'farm_id': farm_id,
                'cooperado': row['cooperado'],
                'codigo': row['codigo'],
                'produto': row['produto'],
                'no_pallet': str(row['no_pallet']),
                'semana': int(row['semana']),
                'data': row['data'],
                'qte': int(row['qte']),
                'area': row['area'],
                'valor_pago': float(row['valor_pago'])
            })
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Processados {len(df)} registros com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 400
    
def processar_vendas_excel(df):
    try:
        # Limpar e preparar os dados
        for index, row in df.iterrows():
            try:
                # Extrair valor numérico da string de valor pago
                valor_pago = float(str(row['valor pago']).replace('R$', '').replace('.', '').replace(',', '.').strip())
                
                # Criar query de inserção
                query = text("""
                    INSERT INTO vendas_uva 
                    (cooperado, codigo, produto, no_pallet, semana, data, qte, area, valor_pago)
                    VALUES 
                    (:cooperado, :codigo, :produto, :no_pallet, :semana, :data, :qte, :area, :valor_pago)
                """)
                
                # Preparar dados para inserção
                dados = {
                    'cooperado': str(row['Cooperado']).strip(),
                    'codigo': str(row['Codigo']).strip(),
                    'produto': str(row['Produto']).strip(),
                    'no_pallet': str(row['No. Pallet']).strip(),
                    'semana': int(row['Semana']),
                    'data': datetime.strptime(str(row['Data']), '%d/%m/%Y').date(),
                    'qte': int(row['Caixas']),
                    'area': str(row['AREA']).strip(),
                    'valor_pago': valor_pago
                }
                
                # Executar a inserção
                db.session.execute(query, dados)
                
            except Exception as e:
                print(f"Erro ao processar linha {index + 1}: {str(e)}")
                print(f"Dados da linha: {row}")
                continue
        
        # Commit das transações
        db.session.commit()
        return True
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro durante o processamento: {str(e)}")
        return False
    
# Adicione estas rotas ao final do arquivo app.py, antes do bloco if __name__ == '__main__':

@app.route('/get_consolidated_report_data')
def get_consolidated_report_data():
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        
        # Métricas gerais consolidadas
        metrics_query = text("""
            WITH FarmCosts AS (
                SELECT 
                    a.farm_id,  # Adicionado alias da tabela
                    SUM(COALESCE(
                        CASE 
                            WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                            ELSE 0 
                        END + COALESCE(a.extra, 0), 0)) as labor_costs
                FROM apontamento a
                JOIN valvulas v ON a.valvula_id = v.id
                JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE a.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                AND v.id != 33
                GROUP BY a.farm_id
            ),
            FuelData AS (
                SELECT 
                    ab.farm_id,  # Adicionado alias da tabela
                    tipo_trator,
                    ABS(SUM(CASE 
                        WHEN quantidade IS NOT NULL THEN quantidade 
                        ELSE 0 
                    END)) as consumo
                FROM abastecimento ab
                WHERE ab.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                AND tipo_trator != 'POSTO DE COMBUSTÍVEL'
                GROUP BY ab.farm_id, tipo_trator
                HAVING consumo > 0
            ),
            ApplicationData AS (
                SELECT 
                    app.farm_id,  # Adicionado alias da tabela
                    COUNT(*) as total_applications
                FROM aplicacoes app
                WHERE app.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                GROUP BY app.farm_id
            ),
            EmployeeData AS (
                SELECT 
                    func.farm_id,  # Adicionado alias da tabela
                    COUNT(*) as active_employees
                FROM funcionarios func
                WHERE func.ativo = 1
                GROUP BY func.farm_id
            )
            SELECT 
                SUM(fc.labor_costs) as total_costs,
                SUM(ad.total_applications) as total_applications,
                (SELECT SUM(active_employees) FROM EmployeeData) as total_employees,
                (SELECT SUM(fd.consumo) FROM FuelData fd) as total_fuel_consumption,
                (SELECT AVG(
                    CASE WHEN a.realizado IS NOT NULL AND a.realizado = 1 THEN 1 ELSE 0 END
                ) * 100 FROM aplicacoes a WHERE a.data BETWEEN :start_date AND :end_date) as avg_efficiency
            FROM FarmCosts fc
            LEFT JOIN ApplicationData ad ON fc.farm_id = ad.farm_id
            LEFT JOIN EmployeeData ed ON fc.farm_id = ed.farm_id
        """)
        
        # Dados por fazenda
        farms_query = text("""
            WITH FarmCosts AS (
                SELECT 
                    a.farm_id,  # Adicionado alias da tabela
                    SUM(COALESCE(
                        CASE 
                            WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                            ELSE 0 
                        END + COALESCE(a.extra, 0), 0)) as labor_costs
                FROM apontamento a
                JOIN valvulas v ON a.valvula_id = v.id
                JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE a.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                AND v.id != 33
                GROUP BY a.farm_id
            ),
            FuelData AS (
                SELECT 
                    ab.farm_id,  # Adicionado alias da tabela
                    ABS(SUM(CASE 
                        WHEN ab.quantidade IS NOT NULL THEN ab.quantidade 
                        ELSE 0 
                    END)) as consumo
                FROM abastecimento ab
                WHERE ab.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                AND ab.tipo_trator != 'POSTO DE COMBUSTÍVEL'
                GROUP BY ab.farm_id
            ),
            ApplicationData AS (
                SELECT 
                    app.farm_id,  # Adicionado alias da tabela
                    COUNT(*) as total_applications
                FROM aplicacoes app
                WHERE app.data BETWEEN :start_date AND :end_date  # Especificado tabela para data
                GROUP BY app.farm_id
            ),
            EmployeeData AS (
                SELECT 
                    func.farm_id,  # Adicionado alias da tabela
                    COUNT(*) as active_employees
                FROM funcionarios func
                WHERE func.ativo = 1
                GROUP BY func.farm_id
            ),
            StockData AS (
                SELECT 
                    reg.farm_id,
                    SUM(
                        CASE 
                            WHEN reg.tipo_movimento = 'ENTRADA' THEN reg.quantidade * COALESCE(reg.valor_unitario, 0)
                            ELSE 0
                        END
                    ) as estoque_valor
                FROM registro_estoque reg
                JOIN produtos p ON reg.produto_id = p.id
                WHERE reg.data <= :end_date
                GROUP BY reg.farm_id
            )
            SELECT 
                fz.id,
                fz.nome,
                COALESCE(fc.labor_costs, 0) as custos,
                COALESCE(ad.total_applications, 0) as aplicacoes,
                COALESCE(ed.active_employees, 0) as funcionarios,
                COALESCE(fd.consumo, 0) as combustivel,
                COALESCE(sd.estoque_valor, 0) as estoque
            FROM fazendas fz
            LEFT JOIN FarmCosts fc ON fz.id = fc.farm_id
            LEFT JOIN ApplicationData ad ON fz.id = ad.farm_id
            LEFT JOIN EmployeeData ed ON fz.id = ed.farm_id
            LEFT JOIN FuelData fd ON fz.id = fd.farm_id
            LEFT JOIN StockData sd ON fz.id = sd.farm_id
        """)
        
        # Dados de consumo de combustível detalhados
        fuel_query = text("""
            SELECT 
                f.nome as farm,
                a.tipo_trator as maquina,
                ABS(SUM(a.quantidade)) as consumo
            FROM abastecimento a
            JOIN fazendas f ON a.farm_id = f.id
            WHERE a.data BETWEEN :start_date AND :end_date
            AND a.tipo_trator != 'POSTO DE COMBUSTÍVEL'
            GROUP BY f.nome, a.tipo_trator
            HAVING consumo > 0
            ORDER BY f.nome, consumo DESC
        """)
        
        # Dados de custos detalhados
        costs_query = text("""
            WITH ValveCosts AS (
                SELECT 
                    fz.nome as farm_name,
                    v.valvula as valve,
                    v.id as valve_id,
                    v.farm_id,
                    COALESCE(SUM(
                        CASE WHEN a.data BETWEEN :start_date AND :end_date 
                        THEN 
                            CASE 
                                WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                                ELSE 0 
                            END + COALESCE(a.extra, 0)
                        ELSE 0 END
                    ), 0) as labor_cost,
                    COALESCE((
                        SELECT SUM(
                            c.quantidade * COALESCE(
                                (SELECT valor_unitario 
                                FROM registro_estoque re 
                                WHERE re.produto_id = c.produto_id 
                                AND re.tipo_movimento = 'ENTRADA'
                                AND re.data <= c.data
                                ORDER BY re.data DESC, re.id DESC
                                LIMIT 1),
                                0
                            ) / 1000
                        )
                        FROM calda c
                        JOIN aplicacoes a ON c.aplicacao_id = a.id
                        WHERE a.setor = v.id
                        AND a.realizado = 1
                        AND a.data BETWEEN :start_date AND :end_date
                        AND a.farm_id = v.farm_id
                        AND c.tipo_movimento = 'SAIDA'
                    ), 0) as inputs_cost
                FROM valvulas v
                JOIN fazendas fz ON v.farm_id = fz.id
                LEFT JOIN apontamento a ON v.id = a.valvula_id
                LEFT JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE v.id != 33
                GROUP BY fz.nome, v.valvula, v.id, v.farm_id
            )
            SELECT 
                farm_name,
                valve,
                labor_cost,
                inputs_cost,
                (labor_cost + inputs_cost) as total_cost
            FROM ValveCosts
            ORDER BY farm_name, total_cost DESC
        """)
        
        # KPIs gerais
        kpis_query = text("""
            WITH AreaTotal AS (
                SELECT 
                    farm_id,
                    SUM(COALESCE(area_hectare, 0)) as area_total
                FROM valvulas
                WHERE id != 33
                GROUP BY farm_id
            ),
            FuelConsumption AS (
                SELECT 
                    farm_id,
                    ABS(SUM(CASE 
                        WHEN quantidade IS NOT NULL THEN quantidade 
                        ELSE 0 
                    END)) as consumo_total
                FROM abastecimento
                WHERE data BETWEEN :start_date AND :end_date
                AND tipo_trator != 'POSTO DE COMBUSTÍVEL'
                GROUP BY farm_id
            ),
            CostData AS (
                SELECT 
                    a.farm_id,
                    SUM(COALESCE(
                        CASE 
                            WHEN f.tipo_contratacao = 'AVULSO' THEN 80
                            ELSE 0 
                        END + COALESCE(extra, 0), 0)) as labor_costs
                FROM apontamento a
                JOIN valvulas v ON a.valvula_id = v.id
                JOIN funcionarios f ON a.funcionario_id = f.id
                WHERE data BETWEEN :start_date AND :end_date
                AND v.id != 33
                GROUP BY a.farm_id
            )
            SELECT 
                SUM(at.area_total) as total_area,
                SUM(fc.consumo_total) as total_fuel,
                SUM(cd.labor_costs) as total_costs,
                CASE 
                    WHEN SUM(at.area_total) > 0 THEN SUM(cd.labor_costs) / SUM(at.area_total)
                    ELSE 0
                END as avg_cost_per_hectare,
                CASE 
                    WHEN SUM(at.area_total) > 0 THEN SUM(fc.consumo_total) / SUM(at.area_total)
                    ELSE 0
                END as avg_fuel_consumption
            FROM AreaTotal at
            JOIN FuelConsumption fc ON at.farm_id = fc.farm_id
            JOIN CostData cd ON at.farm_id = cd.farm_id
        """)
        
        # Executar as consultas
        metrics_data = db.session.execute(metrics_query, {
            'start_date': start_date,
            'end_date': end_date
        }).first()
        
        farms_data = db.session.execute(farms_query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()
        
        fuel_data = db.session.execute(fuel_query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()
        
        costs_data = db.session.execute(costs_query, {
            'start_date': start_date,
            'end_date': end_date
        }).fetchall()
        
        kpis_data = db.session.execute(kpis_query, {
            'start_date': start_date,
            'end_date': end_date
        }).first()
        
        # Preparar o resultado
        return jsonify({
            'status': 'success',
            'metrics': {
                'totalCosts': float(metrics_data.total_costs or 0),
                'totalFuelConsumption': float(metrics_data.total_fuel_consumption or 0),
                'totalApplications': int(metrics_data.total_applications or 0),
                'totalEmployees': int(metrics_data.total_employees or 0),
                'avgEfficiency': float(metrics_data.avg_efficiency or 0)
            },
            'farms': [{
                'id': row[0], 
                'nome': row.nome,
                'custos': float(row.custos or 0),
                'aplicacoes': int(row.aplicacoes or 0),
                'funcionarios': int(row.funcionarios or 0),
                'combustivel': float(row.combustivel or 0),
                'estoque': float(row.estoque or 0)
            } for row in farms_data] if farms_data else [],
            # Garantir valores padrão para todas as propriedades
            'fuelData': [{
                'farm': row.farm,
                'maquina': row.maquina,
                'consumo': float(row.consumo or 0)
            } for row in fuel_data] if fuel_data else [],

            'costsDetails': [{
                'farm_name': row.farm_name,
                'valve': row.valve,
                'laborCost': float(row.labor_cost or 0),
                'inputsCost': float(row.inputs_cost or 0),
                'totalCost': float(row.total_cost or 0)
            } for row in costs_data] if costs_data else [],

            'kpis': {
                'avgCostPerHectare': float(kpis_data.avg_cost_per_hectare or 0),
                'avgFuelConsumption': float(kpis_data.avg_fuel_consumption or 0),
                'totalArea': float(kpis_data.total_area or 0),
                'totalFuel': float(kpis_data.total_fuel or 0)
            } if kpis_data else {'avgCostPerHectare': 0, 'avgFuelConsumption': 0, 'totalArea': 0, 'totalFuel': 0}
        })
    except Exception as e:
        # Melhorar o log de erros
        print(f"Erro detalhado ao carregar dados para relatório consolidado: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/get_consolidated_stock')
def get_consolidated_stock():
    try:
        query = text("""
            WITH UltimoValor AS (
                SELECT 
                    re.farm_id,  # Especificando a tabela (re.)
                    re.produto_id,
                    MAX(re.id) as ultimo_id
                FROM registro_estoque re
                WHERE re.tipo_movimento = 'ENTRADA'
                AND re.valor_unitario IS NOT NULL
                GROUP BY re.farm_id, re.produto_id
            ),
            EstoqueRegular AS (
                SELECT 
                    re.farm_id,
                    p.id,
                    SUM(CASE 
                        WHEN re.tipo_movimento = 'ENTRADA' THEN re.quantidade 
                        WHEN re.tipo_movimento = 'SAIDA' THEN -re.quantidade 
                        ELSE 0 
                    END) as saldo_atual
                FROM produtos p
                LEFT JOIN registro_estoque re ON p.id = re.produto_id
                WHERE p.ativo = 1
                GROUP BY re.farm_id, p.id
                HAVING saldo_atual > 0
            ),
            EstoqueTotal AS (
                SELECT 
                    er.farm_id,  # Especificando a tabela (er.)
                    SUM(
                        er.saldo_atual * COALESCE(
                            (SELECT re2.valor_unitario 
                            FROM registro_estoque re2 
                            WHERE re2.id = uv.ultimo_id),
                            0
                        )
                    ) as valor_total
                FROM EstoqueRegular er
                LEFT JOIN UltimoValor uv ON er.farm_id = uv.farm_id AND er.id = uv.produto_id
                GROUP BY er.farm_id
                
                UNION ALL
                
                SELECT 
                    ec.farm_id,  # Especificando a tabela (ec.)
                    SUM(
                        ec.saldo_atual * COALESCE(
                            (SELECT re3.valor_unitario 
                            FROM registro_estoque re3 
                            WHERE re3.farm_id = ec.farm_id
                            AND re3.produto_id = ec.id
                            AND re3.tipo_movimento = 'ENTRADA'
                            AND re3.valor_unitario IS NOT NULL
                            ORDER BY re3.data DESC, re3.id DESC
                            LIMIT 1),
                            0
                        )
                    ) as valor_total
                FROM EstoqueCalda ec
                GROUP BY ec.farm_id
            )
            SELECT SUM(et.valor_total) as valor_total
            FROM EstoqueTotal et
        """)
        
        result = db.session.execute(query).scalar()
        
        return jsonify({
            'status': 'success',
            'valor_total': float(result) if result else 0
        })
    except Exception as e:
        print(f"Erro ao buscar valor total consolidado do estoque: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/get_consolidated_vendas')
def get_consolidated_vendas():
    try:
        ano = request.args.get('ano', '2025')
        
        # Query para vendas por mês e fazenda
        vendas_query = text("""
            SELECT 
                f.nome as farm_name,
                DATE_FORMAT(vu.data, '%m/%Y') as mes,
                SUM(vu.qte * 5) as quilo,
                SUM(vu.valor_pago) as valor
            FROM vendas_uva vu
            JOIN fazendas f ON vu.farm_id = f.id
            WHERE YEAR(vu.data) = :ano
            GROUP BY f.nome, mes
            ORDER BY f.nome, STR_TO_DATE(mes, '%m/%Y')
        """)
        
        # Query para resumo por fazenda
        resumo_query = text("""
            SELECT 
                f.nome,
                SUM(vu.qte * 5) as total_kg,
                SUM(vu.valor_pago) as valor_total,
                CASE 
                    WHEN SUM(vu.qte * 5) > 0 THEN SUM(vu.valor_pago) / SUM(vu.qte * 5)
                    ELSE 0
                END as preco_medio
            FROM vendas_uva vu
            JOIN fazendas f ON vu.farm_id = f.id
            WHERE YEAR(vu.data) = :ano
            GROUP BY f.nome
            ORDER BY total_kg DESC
        """)
        
        vendas_result = db.session.execute(vendas_query, {'ano': ano}).fetchall()
        resumo_result = db.session.execute(resumo_query, {'ano': ano}).fetchall()
        
        # Calcular estatísticas gerais
        total_kg = sum(float(row.total_kg) for row in resumo_result)
        total_valor = sum(float(row.valor_total) for row in resumo_result)
        avg_price = total_valor / total_kg if total_kg > 0 else 0
        
        return jsonify({
            'status': 'success',
            'vendas': [{
                'farm_name': row.farm_name,
                'mes': row.mes,
                'quilo': float(row.quilo),
                'valor': float(row.valor)
            } for row in vendas_result],
            'vendas_por_fazenda': [{
                'nome': row.nome,
                'total_kg': float(row.total_kg),
                'valor_total': float(row.valor_total),
                'preco_medio': float(row.preco_medio)
            } for row in resumo_result],
            'stats': {
                'total_kg': total_kg,
                'total_valor': total_valor,
                'avg_price': avg_price
            }
        })
    except Exception as e:
        print(f"Erro ao carregar dados consolidados de vendas: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    
@app.route('/get_valvulas_poda')
def get_valvulas_poda():
    try:
        farm_id = session.get('farm_id')
        query = text("""
            SELECT id, valvula, variedade, data_poda, area_hectare
            FROM valvulas 
            WHERE farm_id = :farm_id
            ORDER BY valvula ASC
        """)
        
        result = db.session.execute(query, {'farm_id': farm_id})
        valvulas = [{
            'id': row.id,
            'valvula': row.valvula,
            'variedade': row.variedade,
            'data_poda': row.data_poda.isoformat() if row.data_poda else None,
            'area_hectare': float(row.area_hectare) if row.area_hectare else None
        } for row in result]
        
        return jsonify({
            'status': 'success',
            'valvulas': valvulas
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Erro ao buscar válvulas: {str(e)}'
        }), 400

@app.route('/update_poda_date/<int:id>', methods=['PUT'])
def update_poda_date(id):
    try:
        farm_id = session.get('farm_id')
        data = request.get_json()
        data_poda = data.get('data_poda')
        
        # Garantir que a data seja tratada como uma data sem fuso horário
        if data_poda:
            data_poda_obj = datetime.strptime(data_poda, '%Y-%m-%d').date()
        else:
            data_poda_obj = None
        
        query = text("""
            UPDATE valvulas 
            SET data_poda = :data_poda
            WHERE id = :id
            AND farm_id = :farm_id
        """)
        
        db.session.execute(query, {
            'id': id,
            'data_poda': data_poda_obj,
            'farm_id': farm_id
        })
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Data de poda atualizada com sucesso!'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Erro ao atualizar data de poda: {str(e)}'
        }), 400

@app.route('/download_podas_excel')
def download_podas_excel():
    try:
        farm_id = session.get('farm_id')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Datas de Poda"
        
        # Buscar dados
        query = text("""
            SELECT 
                valvula as "Válvula",
                variedade as "Variedade",
                DATE_FORMAT(data_poda, '%d/%m/%Y') as "Data da Poda",
                DATEDIFF(CURRENT_DATE, data_poda) as "DAP",
                area_hectare as "Área (ha)"
            FROM valvulas
            WHERE farm_id = :farm_id
            ORDER BY valvula
        """)
        
        df = pd.read_sql(query, db.engine, params={'farm_id': farm_id})
        
        # Converter DataFrame para Excel
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        style_excel_worksheet(ws)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"datas_poda_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        print(f"Erro ao gerar Excel de podas: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
if __name__ == '__main__':
    app.run(debug=True)
